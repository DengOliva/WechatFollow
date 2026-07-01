from __future__ import annotations

import html
import json
import mimetypes
import os
import secrets
import shutil
import sqlite3
import subprocess
import tempfile
import threading
import time
import urllib.parse
from datetime import date, datetime
from email.parser import BytesParser
from email.policy import default
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("APP_DATA_DIR", ROOT))
DB_PATH = DATA_DIR / "tasks.db"
UIA_SCRIPT = ROOT / "wechat_uia.ps1"
UPLOAD_DIR = DATA_DIR / "data" / "uploads"
HOST = "127.0.0.1"
PORT = int(os.environ.get("APP_PORT", "8000"))
MAX_UPLOAD_SIZE = 20 * 1024 * 1024
PUBLIC_BASE_URL = os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")
NORMAL_REMINDER_SLOTS = ("08:30",)
URGENT_REMINDER_SLOTS = ("08:30", "14:30")
DEFAULT_ADMIN_RECIPIENT = "邓宇聪"
WECHAT_LOCK = threading.Lock()


def db() -> sqlite3.Connection:
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def initialize_database() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    with db() as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                content TEXT NOT NULL,
                deadline TEXT NOT NULL,
                assignee TEXT NOT NULL,
                created_at TEXT NOT NULL,
                notified_at TEXT,
                submit_token TEXT
            )
            """
        )
        columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(tasks)").fetchall()
        }
        if "submit_token" not in columns:
            connection.execute("ALTER TABLE tasks ADD COLUMN submit_token TEXT")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
            """
        )
        connection.execute(
            "INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)",
            ("admin_recipient", DEFAULT_ADMIN_RECIPIENT),
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS submissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER NOT NULL,
                note TEXT NOT NULL,
                original_filename TEXT NOT NULL,
                stored_filename TEXT NOT NULL,
                submitted_at TEXT NOT NULL,
                FOREIGN KEY (task_id) REFERENCES tasks(id)
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS reminder_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER NOT NULL,
                reminder_date TEXT NOT NULL,
                reminder_slot TEXT NOT NULL DEFAULT '08:30',
                attempted_at TEXT NOT NULL,
                success INTEGER NOT NULL,
                detail TEXT NOT NULL,
                UNIQUE(task_id, reminder_date, reminder_slot),
                FOREIGN KEY (task_id) REFERENCES tasks(id)
            )
            """
        )
        reminder_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(reminder_logs)").fetchall()
        }
        reminder_indexes = connection.execute(
            "PRAGMA index_list(reminder_logs)"
        ).fetchall()
        has_old_unique = any(
            row["unique"]
            and [
                info["name"]
                for info in connection.execute(
                    f"PRAGMA index_info({row['name']})"
                ).fetchall()
            ]
            == ["task_id", "reminder_date"]
            for row in reminder_indexes
        )
        if "reminder_slot" not in reminder_columns or has_old_unique:
            connection.execute("ALTER TABLE reminder_logs RENAME TO reminder_logs_old")
            connection.execute(
                """
                CREATE TABLE reminder_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id INTEGER NOT NULL,
                    reminder_date TEXT NOT NULL,
                    reminder_slot TEXT NOT NULL DEFAULT '08:30',
                    attempted_at TEXT NOT NULL,
                    success INTEGER NOT NULL,
                    detail TEXT NOT NULL,
                    UNIQUE(task_id, reminder_date, reminder_slot),
                    FOREIGN KEY (task_id) REFERENCES tasks(id)
                )
                """
            )
            old_columns = {
                row["name"]
                for row in connection.execute(
                    "PRAGMA table_info(reminder_logs_old)"
                ).fetchall()
            }
            slot_expr = (
                "reminder_slot"
                if "reminder_slot" in old_columns
                else "'08:30' AS reminder_slot"
            )
            connection.execute(
                f"""
                INSERT OR IGNORE INTO reminder_logs
                    (id, task_id, reminder_date, reminder_slot,
                     attempted_at, success, detail)
                SELECT id, task_id, reminder_date, {slot_expr},
                       attempted_at, success, detail
                FROM reminder_logs_old
                """
            )
            connection.execute("DROP TABLE reminder_logs_old")
        tasks_without_token = connection.execute(
            "SELECT id FROM tasks WHERE submit_token IS NULL OR submit_token = ''"
        ).fetchall()
        for task in tasks_without_token:
            connection.execute(
                "UPDATE tasks SET submit_token = ? WHERE id = ?",
                (secrets.token_urlsafe(24), task["id"]),
            )


def get_tasks() -> list[sqlite3.Row]:
    with db() as connection:
        return connection.execute(
            """
            SELECT tasks.*,
                   (SELECT COUNT(*) FROM submissions
                    WHERE submissions.task_id = tasks.id) AS submission_count,
                   (SELECT MAX(submitted_at) FROM submissions
                    WHERE submissions.task_id = tasks.id) AS last_submitted_at,
                   (SELECT success FROM reminder_logs
                    WHERE reminder_logs.task_id = tasks.id
                    ORDER BY reminder_logs.id DESC LIMIT 1) AS last_reminder_success,
                   (SELECT attempted_at FROM reminder_logs
                    WHERE reminder_logs.task_id = tasks.id
                    ORDER BY reminder_logs.id DESC LIMIT 1) AS last_reminder_at,
                   (SELECT reminder_slot FROM reminder_logs
                    WHERE reminder_logs.task_id = tasks.id
                    ORDER BY reminder_logs.id DESC LIMIT 1) AS last_reminder_slot
            FROM tasks
            ORDER BY tasks.id DESC
            """
        ).fetchall()


def get_task(task_id: int) -> sqlite3.Row | None:
    with db() as connection:
        return connection.execute(
            "SELECT * FROM tasks WHERE id = ?", (task_id,)
        ).fetchone()


def get_task_by_token(token: str) -> sqlite3.Row | None:
    with db() as connection:
        return connection.execute(
            "SELECT * FROM tasks WHERE submit_token = ?", (token,)
        ).fetchone()


def get_setting(key: str, default: str = "") -> str:
    with db() as connection:
        row = connection.execute(
            "SELECT value FROM settings WHERE key = ?", (key,)
        ).fetchone()
    return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    with db() as connection:
        connection.execute(
            """
            INSERT INTO settings (key, value)
            VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (key, value),
        )


def get_admin_recipient() -> str:
    return get_setting("admin_recipient", DEFAULT_ADMIN_RECIPIENT).strip()


def set_admin_recipient(value: str) -> None:
    set_setting("admin_recipient", value.strip() or DEFAULT_ADMIN_RECIPIENT)


def get_submissions(task_id: int) -> list[sqlite3.Row]:
    with db() as connection:
        return connection.execute(
            "SELECT * FROM submissions WHERE task_id = ? ORDER BY id DESC",
            (task_id,),
        ).fetchall()


def get_submission(submission_id: int) -> sqlite3.Row | None:
    with db() as connection:
        return connection.execute(
            "SELECT * FROM submissions WHERE id = ?", (submission_id,)
        ).fetchone()


def delete_task(task_id: int) -> bool:
    submissions = get_submissions(task_id)
    with db() as connection:
        cursor = connection.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
        deleted = cursor.rowcount > 0
        if deleted:
            connection.execute("DELETE FROM submissions WHERE task_id = ?", (task_id,))
            connection.execute("DELETE FROM reminder_logs WHERE task_id = ?", (task_id,))
    if deleted:
        for submission in submissions:
            file_path = UPLOAD_DIR / submission["stored_filename"]
            try:
                if file_path.is_file():
                    file_path.unlink()
            except OSError as exc:
                print(f"[DELETE] 删除上传文件失败：{file_path} {exc}")
    return deleted


def create_task(content: str, deadline: str, assignee: str) -> int:
    with db() as connection:
        cursor = connection.execute(
            """
            INSERT INTO tasks
                (content, deadline, assignee, created_at, submit_token)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                content,
                deadline,
                assignee,
                datetime.now().isoformat(timespec="seconds"),
                secrets.token_urlsafe(24),
            ),
        )
        return int(cursor.lastrowid)


def create_tasks_bulk(tasks: list[tuple[str, str, str]]) -> int:
    count = 0
    with db() as connection:
        for content, deadline, assignee in tasks:
            connection.execute(
                """
                INSERT INTO tasks
                    (content, deadline, assignee, created_at, submit_token)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    content,
                    deadline,
                    assignee,
                    datetime.now().isoformat(timespec="seconds"),
                    secrets.token_urlsafe(24),
                ),
            )
            count += 1
    return count


def mark_notified(task_id: int) -> None:
    with db() as connection:
        connection.execute(
            "UPDATE tasks SET notified_at = ? WHERE id = ?",
            (datetime.now().isoformat(timespec="seconds"), task_id),
        )


def parse_deadline_date(deadline: str) -> date | None:
    text = deadline.strip()
    if not text:
        return None
    candidates = [
        text,
        text.replace("T", " "),
        text.split(" ")[0],
        text.split("T")[0],
    ]
    formats = [
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ]
    for candidate in candidates:
        for fmt in formats:
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                pass
    return None


def is_urgent_or_overdue(task: sqlite3.Row, today: date | None = None) -> bool:
    deadline_date = parse_deadline_date(task["deadline"])
    if deadline_date is None:
        return False
    today = today or date.today()
    days_until_deadline = (deadline_date - today).days
    return days_until_deadline <= 3


def get_pending_reminder_tasks(
    reminder_date: str, reminder_slot: str
) -> list[sqlite3.Row]:
    with db() as connection:
        tasks = connection.execute(
            """
            SELECT tasks.*
            FROM tasks
            WHERE NOT EXISTS (
                SELECT 1 FROM submissions WHERE submissions.task_id = tasks.id
            )
            AND NOT EXISTS (
                SELECT 1 FROM reminder_logs
                WHERE reminder_logs.task_id = tasks.id
                  AND reminder_logs.reminder_date = ?
                  AND reminder_logs.reminder_slot = ?
            )
            ORDER BY tasks.id
            """,
            (reminder_date, reminder_slot),
        ).fetchall()
    if reminder_slot in URGENT_REMINDER_SLOTS and reminder_slot not in NORMAL_REMINDER_SLOTS:
        return [task for task in tasks if is_urgent_or_overdue(task)]
    return tasks


def record_reminder_attempt(
    task_id: int,
    reminder_date: str,
    reminder_slot: str,
    success: bool,
    detail: str,
) -> None:
    with db() as connection:
        connection.execute(
            """
            INSERT OR IGNORE INTO reminder_logs
                (task_id, reminder_date, reminder_slot,
                 attempted_at, success, detail)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                task_id,
                reminder_date,
                reminder_slot,
                datetime.now().isoformat(timespec="seconds"),
                1 if success else 0,
                detail,
            ),
        )


def save_submission(
    task_id: int, note: str, original_filename: str, file_data: bytes
) -> sqlite3.Row:
    safe_name = Path(original_filename).name.strip() or "upload.bin"
    stored_name = f"{secrets.token_hex(16)}{Path(safe_name).suffix[:16]}"
    (UPLOAD_DIR / stored_name).write_bytes(file_data)
    with db() as connection:
        cursor = connection.execute(
            """
            INSERT INTO submissions
                (task_id, note, original_filename, stored_filename, submitted_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                task_id,
                note,
                safe_name,
                stored_name,
                datetime.now().isoformat(timespec="seconds"),
            ),
        )
        return connection.execute(
            "SELECT * FROM submissions WHERE id = ?", (cursor.lastrowid,)
        ).fetchone()


def format_excel_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def normalize_header(value: object) -> str:
    return "".join(format_excel_cell(value).split())


def parse_task_excel(file_data: bytes) -> tuple[list[tuple[str, str, str]], list[str]]:
    try:
        workbook = load_workbook(BytesIO(file_data), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"无法读取 Excel 文件：{exc}") from exc

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Excel 第一行必须是表头。")

    headers = {normalize_header(value): index for index, value in enumerate(header_row)}
    required = ["任务内容及要求", "责任人", "完成期限"]
    required_keys = {name: normalize_header(name) for name in required}
    missing = [name for name, key in required_keys.items() if key not in headers]
    if missing:
        raise ValueError("缺少表头：" + "、".join(missing))

    tasks: list[tuple[str, str, str]] = []
    errors: list[str] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        content = format_excel_cell(row[headers[required_keys["任务内容及要求"]]])
        assignee = format_excel_cell(row[headers[required_keys["责任人"]]])
        deadline = format_excel_cell(row[headers[required_keys["完成期限"]]])

        if not any((content, assignee, deadline)):
            continue

        missing_fields = []
        if not content:
            missing_fields.append("任务内容及要求")
        if not assignee:
            missing_fields.append("责任人")
        if not deadline:
            missing_fields.append("完成期限")
        if missing_fields:
            errors.append(f"第 {row_number} 行缺少：" + "、".join(missing_fields))
            continue

        tasks.append((content, deadline, assignee))

    if not tasks and not errors:
        raise ValueError("Excel 中没有可导入的任务行。")
    return tasks, errors


def send_wechat_message(
    recipient: str, message: str, attachment_path: Path | None = None
) -> tuple[bool, str]:
    command = [
        "powershell.exe",
        "-NoProfile",
        "-STA",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(UIA_SCRIPT),
        "-Recipient",
        recipient,
        "-Message",
        message,
    ]
    if attachment_path is not None:
        command.extend(["-AttachmentPath", str(attachment_path)])
    with WECHAT_LOCK:
        try:
            result = subprocess.run(
                command,
                cwd=ROOT,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=90 if attachment_path is not None else 30,
            )
        except subprocess.TimeoutExpired:
            return False, "操作超时，请确认微信窗口可见且没有被其他弹窗遮挡。"
        except OSError as exc:
            return False, f"无法启动 UI Automation 脚本：{exc}"

    output = (result.stdout or result.stderr).strip()
    messages = {
        "WECHAT_NOT_RUNNING": "未检测到微信进程，请先启动并登录微信桌面端。",
        "WECHAT_WINDOW_NOT_FOUND": "微信正在运行，但找不到主窗口。请从系统托盘打开微信主界面后重试。",
        "CHAT_INPUT_NOT_FOUND": "已打开联系人，但无法定位聊天输入框。请保持微信主窗口完整显示后重试。",
        "ATTACHMENT_NOT_FOUND": "文字已发送，但服务器上的任务完成资料不存在。",
        "SENT": f"已向 {recipient} 发送任务通知。",
        "SENT_WITH_ATTACHMENT": f"已向 {recipient} 发送反馈和任务完成资料。",
    }
    output = messages.get(output, output)
    if result.returncode == 0:
        return True, output or "通知已发送。"
    return False, output or f"发送失败，脚本退出码：{result.returncode}"


def submission_url(task: sqlite3.Row, request_base_url: str) -> str:
    base_url = PUBLIC_BASE_URL or request_base_url
    return f"{base_url}/submit/{task['submit_token']}"


def notification_message(task: sqlite3.Row, link: str) -> str:
    return (
        "+---------- 待办任务 ----------+\n"
        f"| 责任人：{task['assignee']}\n"
        f"| 期限：{task['deadline']}\n"
        "+---------- 任务内容 ----------+\n"
        f"{task['content']}\n"
        "+---------- 提交入口 ----------+\n"
        f"{link}\n"
        "+------------------------------+\n"
        "请在期限前打开链接上传资料。"
    )


def reminder_message(task: sqlite3.Row, link: str, urgent: bool) -> str:
    title = "【临期/逾期任务提醒】" if urgent else "【任务提醒】"
    intro = (
        "这项任务已临近或超过期限，请尽快处理："
        if urgent
        else "你有一项任务尚未提交资料："
    )
    return (
        f"+---------- {title.strip('【】')} ----------+\n"
        f"| {intro}\n"
        f"| 期限：{task['deadline']}\n"
        "+---------- 任务内容 ----------+\n"
        f"{task['content']}\n"
        "+---------- 提交入口 ----------+\n"
        f"{link}\n"
        "+------------------------------+"
    )


def admin_submission_message(
    task: sqlite3.Row, submission: sqlite3.Row
) -> str:
    note = submission["note"].strip() or "未填写"
    return (
        "+---------- 任务完成反馈 ----------+\n"
        f"任务内容：{task['content']}\n"
        f"提交说明：{note}\n"
        f"任务完成资料：{submission['original_filename']}\n"
        "+----------------------------------+"
    )


def send_admin_submission(
    admin_recipient: str, task: sqlite3.Row, submission: sqlite3.Row
) -> tuple[bool, str]:
    stored_path = UPLOAD_DIR / submission["stored_filename"]
    if not stored_path.is_file():
        return False, "服务器上的任务完成资料不存在。"

    original_name = Path(submission["original_filename"]).name or "任务完成资料"
    invalid_chars = '<>:"/\\|?*'
    safe_name = "".join("_" if char in invalid_chars else char for char in original_name)
    safe_name = safe_name.rstrip(" .") or "任务完成资料"
    with tempfile.TemporaryDirectory(prefix="wechat-follow-submit-") as temp_dir:
        outgoing_path = Path(temp_dir) / safe_name
        shutil.copy2(stored_path, outgoing_path)
        return send_wechat_message(
            admin_recipient,
            admin_submission_message(task, submission),
            outgoing_path,
        )


def run_reminder_slot(reminder_slot: str) -> None:
    reminder_date = date.today().isoformat()
    tasks = get_pending_reminder_tasks(reminder_date, reminder_slot)
    urgent_slot = reminder_slot in URGENT_REMINDER_SLOTS and reminder_slot not in NORMAL_REMINDER_SLOTS
    print(f"[REMINDER] {reminder_date} {reminder_slot} 待提醒任务：{len(tasks)}")
    for task in tasks:
        link = submission_url(task, PUBLIC_BASE_URL or f"http://{HOST}:{PORT}")
        ok, detail = send_wechat_message(
            task["assignee"],
            reminder_message(task, link, urgent_slot or is_urgent_or_overdue(task)),
        )
        record_reminder_attempt(task["id"], reminder_date, reminder_slot, ok, detail)
        print(
            f"[REMINDER] slot={reminder_slot} task={task['id']} "
            f"success={ok} detail={detail}"
        )
        time.sleep(1)


def reminder_loop() -> None:
    reminder_slots = tuple(dict.fromkeys(NORMAL_REMINDER_SLOTS + URGENT_REMINDER_SLOTS))
    last_run_keys: set[str] = set()
    while True:
        now = datetime.now()
        today = now.date().isoformat()
        current_time = now.strftime("%H:%M")
        for slot in reminder_slots:
            key = f"{today} {slot}"
            if current_time == slot and key not in last_run_keys:
                run_reminder_slot(slot)
                last_run_keys.add(key)
        time.sleep(30)


def common_styles() -> str:
    return """
    :root { font-family: "Microsoft YaHei", sans-serif; color: #17202a; }
    * { box-sizing: border-box; }
    body { margin: 0; background: #f3f6f5; }
    main { max-width: 1120px; margin: 36px auto; padding: 0 20px 48px; }
    h1 { margin: 0 0 8px; letter-spacing: -.02em; }
    h2 { margin: 0 0 18px; font-size: 19px; }
    .hint { color: #667085; margin-top: 0; }
    .card { background: white; border: 1px solid #e4e9e7; border-radius: 16px;
            padding: 24px; margin: 20px 0; box-shadow: 0 8px 28px rgba(20,52,39,.06); }
    .grid { display: grid; grid-template-columns: 2fr 1fr 1fr auto; gap: 12px;
            align-items: end; }
    .import-grid { display: grid; grid-template-columns: 1fr auto; gap: 12px;
            align-items: end; }
    label { display: grid; gap: 7px; font-weight: 600; }
    input, textarea { box-sizing: border-box; width: 100%; padding: 10px 12px;
            border: 1px solid #cfd8d5; border-radius: 9px; font: inherit; background: white; }
    input:focus, textarea:focus { outline: 3px solid rgba(7,193,96,.13);
            border-color: #07a954; }
    textarea { min-height: 100px; resize: vertical; }
    button, .button { border: 0; border-radius: 9px; padding: 11px 18px;
            background: #07c160; color: white; font-weight: 700; cursor: pointer;
            white-space: nowrap; text-decoration: none; display: inline-flex;
            align-items: center; justify-content: center; }
    button.secondary, .button.secondary { background: #eef8f2; color: #087d42;
            padding: 8px 12px; }
    button.outline, .button.outline { background: white; color: #087d42;
            border: 1px solid #b9ddc9; }
    button.danger { background: #fff1f0; color: #b42318; box-shadow: none; }
    table { width: 100%; border-collapse: collapse; }
    th, td { padding: 12px 10px; border-bottom: 1px solid #edf0ef; text-align: left;
            vertical-align: top; }
    th { color: #667085; font-size: 14px; }
    .notice { padding: 12px 16px; border-radius: 8px; background: #eaf9f0;
            color: #087d42; }
    .notice.error { background: #fff0f0; color: #b42318; }
    .status { display: inline-flex; padding: 5px 9px; border-radius: 999px;
            font-size: 13px; font-weight: 700; color: #087d42; background: #eaf9f0; }
    .status.pending { color: #9a5b00; background: #fff6e5; }
    .actions { display: flex; gap: 8px; flex-wrap: wrap; }
    .submit-link { margin-top: 8px; max-width: 360px; overflow-wrap: anywhere;
            color: #667085; font-size: 12px; }
    .task-list { display: grid; gap: 16px; margin-top: 20px; }
    .task-card { background: white; border: 1px solid #e4e9e7; border-radius: 16px;
            padding: 22px; box-shadow: 0 8px 28px rgba(20,52,39,.05); }
    .task-head { display: flex; justify-content: space-between; gap: 18px;
            align-items: flex-start; }
    .task-id { color: #98a2b3; font-size: 13px; margin-bottom: 5px; }
    .task-title { margin: 0; font-size: 18px; line-height: 1.55; }
    .task-meta { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 10px; margin: 18px 0; }
    .meta-item { background: #f7f9f8; border-radius: 10px; padding: 12px; }
    .meta-label { display: block; color: #667085; font-size: 12px; margin-bottom: 5px; }
    .meta-value { font-weight: 650; overflow-wrap: anywhere; }
    .entry-box { border-top: 1px solid #edf0ef; padding-top: 16px; }
    .entry-label { font-size: 13px; color: #667085; margin-bottom: 8px; }
    .entry-actions { display: flex; gap: 9px; flex-wrap: wrap; }
    .import-tip { color: var(--muted); font-size: 13px; margin: 9px 0 0; }
    .upload-shell { max-width: 680px; }
    .submit-hero { background: linear-gradient(135deg, #063d25, #078848);
            color: white; border-radius: 20px; padding: 28px; margin-bottom: 18px; }
    .submit-hero .hint { color: rgba(255,255,255,.75); margin-bottom: 0; }
    .task-summary { display: grid; gap: 15px; }
    .summary-label { color: #667085; font-size: 13px; display: block; margin-bottom: 5px; }
    .summary-value { font-size: 17px; line-height: 1.65; font-weight: 650; }
    .file-field { border: 1px dashed #9fcbb2; border-radius: 12px; padding: 16px;
            background: #f6fbf8; }
    .submit-button { width: 100%; padding: 14px 20px; font-size: 16px; }
    .privacy-note { text-align: center; color: #98a2b3; font-size: 12px; }
    form { margin: 0; }
    @media (max-width: 800px) {
      .grid { grid-template-columns: 1fr; }
      .import-grid { grid-template-columns: 1fr; }
      .table-wrap { overflow-x: auto; }
      main { margin-top: 22px; padding: 0 14px 36px; }
      .card, .task-card { padding: 18px; border-radius: 14px; }
      .task-head { display: block; }
      .task-head .status { margin-top: 10px; }
      .task-meta { grid-template-columns: 1fr; gap: 8px; }
      .entry-actions { display: grid; grid-template-columns: 1fr 1fr; }
      .entry-actions form, .entry-actions button, .entry-actions .button { width: 100%; }
      .submit-hero { border-radius: 16px; padding: 22px; }
    }
    """


def premium_styles() -> str:
    return """
    :root { --green:#08a958; --green-dark:#075d35; --line:#e3ebe7; --muted:#6b7a72; }
    body { background:
      radial-gradient(circle at 8% 0%, rgba(8,169,88,.11), transparent 26rem),
      #f4f7f5; color:#18231e; }
    main { max-width:1160px; margin:0 auto; padding:30px 22px 56px; }
    .topbar { display:flex; align-items:center; justify-content:space-between;
      gap:20px; padding:5px 2px 24px; }
    .brand { display:flex; align-items:center; gap:13px; }
    .brand-mark { width:46px; height:46px; border-radius:14px; display:grid;
      place-items:center; color:#fff; font-size:21px; font-weight:800;
      background:linear-gradient(145deg,#08c769,#067c42);
      box-shadow:0 10px 24px rgba(7,153,79,.25); }
    .brand-title { font-size:22px; font-weight:800; letter-spacing:-.02em; }
    .brand-subtitle { color:var(--muted); font-size:13px; margin-top:3px; }
    .schedule-chip { padding:9px 13px; border:1px solid #cce7d8;
      border-radius:999px; color:var(--green-dark); background:#f5fcf8;
      font-size:13px; font-weight:700; }
    .stats { display:grid; grid-template-columns:repeat(3,1fr); gap:14px;
      margin-bottom:18px; }
    .stat { background:rgba(255,255,255,.92); border:1px solid var(--line);
      border-radius:16px; padding:18px 20px;
      box-shadow:0 8px 24px rgba(26,56,42,.05); }
    .stat-label { color:var(--muted); font-size:13px; }
    .stat-value { margin-top:7px; font-size:28px; font-weight:800; line-height:1; }
    .stat-note { color:#8a9891; font-size:12px; margin-top:8px; }
    .card { border-color:var(--line); border-radius:18px;
      box-shadow:0 12px 35px rgba(20,52,39,.065); }
    .section-head { display:flex; justify-content:space-between; align-items:flex-end;
      gap:15px; margin-bottom:17px; }
    .section-head h2 { margin:0; }
    .section-note { color:var(--muted); font-size:13px; }
    input,textarea { padding:12px 13px; border-radius:11px; }
    button,.button { border-radius:10px; background:linear-gradient(135deg,#08bd63,#07944f);
      box-shadow:0 5px 13px rgba(7,153,79,.15);
      transition:transform .14s,filter .14s; }
    button:hover,.button:hover { transform:translateY(-1px); filter:brightness(.98); }
    button.secondary,.button.secondary,button.outline,.button.outline { box-shadow:none; }
    .notice { border:1px solid #ccebd8; border-radius:12px; margin-bottom:18px;
      font-weight:650; }
    .task-list { gap:17px; margin-top:16px; }
    .task-card { position:relative; overflow:hidden; border-color:var(--line);
      border-radius:18px; padding:23px; transition:transform .16s,box-shadow .16s; }
    .task-card:before { content:""; position:absolute; inset:0 auto 0 0; width:4px;
      background:linear-gradient(#09bf65,#078448); }
    .task-card:hover { transform:translateY(-2px);
      box-shadow:0 14px 36px rgba(20,52,39,.09); }
    .task-id { font-size:12px; font-weight:700; letter-spacing:.08em; }
    .task-title { font-size:19px; }
    .meta-item { border:1px solid #edf2ef; border-radius:12px; padding:13px; }
    .entry-label { font-weight:700; }
    .submit-link { max-width:620px; color:#849089; margin-top:11px; }
    .upload-shell { max-width:720px; }
    .submit-hero { position:relative; overflow:hidden;
      background:linear-gradient(135deg,#063d25 0%,#078848 72%,#08ad5a 100%);
      border-radius:24px; padding:34px 32px;
      box-shadow:0 18px 40px rgba(5,93,51,.22); }
    .submit-hero:after { content:""; position:absolute; width:220px; height:220px;
      border-radius:50%; right:-85px; top:-105px; background:rgba(255,255,255,.10); }
    .submit-kicker { font-size:12px; font-weight:800; letter-spacing:.14em;
      color:rgba(255,255,255,.72); margin-bottom:10px; }
    .deadline-box { background:#fff8e9; border:1px solid #f2dfb3;
      border-radius:12px; padding:13px 15px; }
    .file-field { border:1.5px dashed #8fc8a8; border-radius:15px; padding:20px;
      background:linear-gradient(180deg,#f8fcfa,#f2faf5); }
    .steps { display:flex; align-items:center; gap:8px; margin:2px 0 18px; }
    .step { flex:1; height:4px; background:#dce7e1; border-radius:999px; }
    .step.active { background:var(--green); }
    .workspace-grid { display:grid; grid-template-columns:minmax(0,1.45fr) minmax(310px,.75fr);
      gap:18px; align-items:start; margin:20px 0 26px; }
    .toolbar { display:flex; justify-content:space-between; align-items:center; gap:14px;
      background:rgba(255,255,255,.86); border:1px solid var(--line);
      border-radius:16px; padding:12px; margin-bottom:18px; }
    .tabs { display:flex; gap:8px; flex-wrap:wrap; }
    .tab { display:inline-flex; align-items:center; padding:10px 14px; border-radius:11px;
      color:#476056; text-decoration:none; font-weight:700; background:#f3f7f5; }
    .tab.active { color:white; background:linear-gradient(135deg,#08bd63,#07944f);
      box-shadow:0 5px 13px rgba(7,153,79,.15); }
    .side-stack { display:grid; gap:16px; }
    .workspace-grid .card { margin:0; }
    .compact-card { padding:21px; }
    .compact-card .section-head { display:block; margin-bottom:14px; }
    .compact-card .section-note { margin-top:5px; line-height:1.6; }
    .create-card { min-height:100%; }
    .create-card .grid { grid-template-columns:1fr 1fr; }
    .create-card .grid label:first-child { grid-column:1 / -1; }
    .create-card .grid button { align-self:end; }
    .entry-actions { align-items:center; justify-content:space-between; }
    .entry-primary,.entry-secondary { display:flex; gap:9px; flex-wrap:wrap; }
    .entry-secondary { justify-content:flex-end; }
    .task-list-title { margin-top:8px; }
    @media (max-width:800px) {
      main { padding:20px 14px 36px; }
      .topbar { align-items:flex-start; }
      .schedule-chip { display:none; }
      .stats { grid-template-columns:1fr 1fr; }
      .stat:first-child { grid-column:1 / -1; }
      .submit-hero { border-radius:18px; padding:27px 22px; }
      .workspace-grid { grid-template-columns:1fr; }
      .toolbar { display:grid; grid-template-columns:1fr; }
      .toolbar form button { width:100%; }
      .create-card .grid { grid-template-columns:1fr; }
      .create-card .grid label:first-child { grid-column:auto; }
      .entry-actions { display:grid; grid-template-columns:1fr; }
      .entry-primary,.entry-secondary { display:grid; grid-template-columns:1fr 1fr; }
      .entry-primary a,.entry-primary button,.entry-secondary a,.entry-secondary button,
      .entry-secondary form { width:100%; }
    }
    """


def page(
    request_base_url: str,
    view: str = "pending",
    message: str = "",
    error: bool = False,
) -> bytes:
    tasks = get_tasks()
    total_count = len(tasks)
    submitted_count = sum(1 for task in tasks if task["submission_count"])
    pending_count = total_count - submitted_count
    view = "done" if view == "done" else "pending"
    visible_tasks = [
        task
        for task in tasks
        if (task["submission_count"] > 0) == (view == "done")
    ]
    admin_recipient = get_admin_recipient()
    cards = []
    for task in visible_tasks:
        notification_status = (
            f"已通知<br><small>{html.escape(task['notified_at'].replace('T', ' '))}</small>"
            if task["notified_at"]
            else "尚未通知"
        )
        if task["last_reminder_at"]:
            reminder_result = "成功" if task["last_reminder_success"] else "失败"
            reminder_status = (
                f"{reminder_result}<br><small>"
                f"{html.escape(task['last_reminder_at'].replace('T', ' '))}"
                f" {html.escape(task['last_reminder_slot'] or '')}</small>"
            )
        else:
            reminder_status = "尚未自动提醒"
        submit_status = (
            f'<span class="status">已提交 {task["submission_count"]} 次</span>'
            if task["submission_count"]
            else '<span class="status pending">等待提交</span>'
        )
        link = submission_url(task, request_base_url)
        escaped_link = html.escape(link)
        cards.append(
            f"""
            <article class="task-card">
              <div class="task-head">
                <div>
                  <div class="task-id">任务 #{task['id']}</div>
                  <h3 class="task-title">{html.escape(task['content'])}</h3>
                </div>
                {submit_status}
              </div>
              <div class="task-meta">
                <div class="meta-item">
                  <span class="meta-label">责任人</span>
                  <span class="meta-value">{html.escape(task['assignee'])}</span>
                </div>
                <div class="meta-item">
                  <span class="meta-label">提交期限</span>
                  <span class="meta-value">{html.escape(task['deadline'])}</span>
                </div>
                <div class="meta-item">
                  <span class="meta-label">通知状态</span>
                  <span class="meta-value">{notification_status}</span>
                </div>
                <div class="meta-item">
                  <span class="meta-label">自动提醒</span>
                  <span class="meta-value">{reminder_status}</span>
                </div>
              </div>
              <div class="entry-box">
                <div class="entry-label">责任人资料提交入口</div>
                <div class="entry-actions">
                  <div class="entry-primary">
                    <a class="button" href="{escaped_link}" target="_blank">打开提交页面</a>
                    <button class="outline" type="button"
                      data-link="{html.escape(link, quote=True)}"
                      onclick="copySubmissionLink(this)">复制提交链接</button>
                  </div>
                  <div class="entry-secondary">
                    <form method="post" action="/tasks/{task['id']}/notify">
                      <button class="secondary" type="submit">发送微信通知</button>
                    </form>
                    <a class="button secondary" href="/tasks/{task['id']}">查看提交资料</a>
                    <form method="post" action="/tasks/{task['id']}/delete"
                      onsubmit="return confirm('确定删除这个任务吗？此操作会同时删除提交记录和上传文件。')">
                      <button class="danger" type="submit">删除任务</button>
                    </form>
                  </div>
                </div>
                <div class="submit-link">{escaped_link}</div>
              </div>
            </article>
            """
        )

    notice = ""
    if message:
        css_class = "notice error" if error else "notice"
        notice = f'<div class="{css_class}">{html.escape(message)}</div>'

    body = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>微信任务跟踪</title>
  <style>{common_styles()}{premium_styles()}</style>
</head>
<body>
<main>
  <header class="topbar">
    <div class="brand">
      <div class="brand-mark">任</div>
      <div>
        <div class="brand-title">任务跟踪中心</div>
        <div class="brand-subtitle">微信通知、资料提交与进度管理</div>
      </div>
    </div>
    <div class="schedule-chip">08:30 常规提醒，14:30 临期/逾期加提醒</div>
  </header>
  <section class="stats">
    <div class="stat">
      <div class="stat-label">全部任务</div>
      <div class="stat-value">{total_count}</div>
      <div class="stat-note">当前已创建任务总数</div>
    </div>
    <div class="stat">
      <div class="stat-label">等待提交</div>
      <div class="stat-value">{pending_count}</div>
      <div class="stat-note">临期或逾期会增加下午提醒</div>
    </div>
    <div class="stat">
      <div class="stat-label">已提交</div>
      <div class="stat-value">{submitted_count}</div>
      <div class="stat-note">至少提交过一次资料</div>
    </div>
  </section>
  <section class="toolbar">
    <div class="tabs">
      <a class="tab {'active' if view == 'pending' else ''}" href="/?view=pending">待处理 {pending_count}</a>
      <a class="tab {'active' if view == 'done' else ''}" href="/?view=done">已完成 {submitted_count}</a>
    </div>
    <form method="post" action="/tasks/remind-all"
      onsubmit="return confirm('确定提醒当前所有待处理任务的责任人吗？')">
      <button type="submit">一键提醒待处理任务</button>
    </form>
  </section>
  {notice}
  <section class="workspace-grid">
    <div class="card create-card">
      <div class="section-head">
        <h2>创建新任务</h2>
        <div class="section-note">录入后自动生成专属资料提交入口，可立即发送微信通知。</div>
      </div>
      <form method="post" action="/tasks" class="grid">
        <label>任务内容
          <input name="content" required maxlength="500" placeholder="例如：提交本周项目进度">
        </label>
        <label>任务期限
          <input name="deadline" type="datetime-local" required>
        </label>
        <label>责任人（微信昵称）
          <input name="assignee" required maxlength="100" placeholder="需能在微信中搜索">
        </label>
        <button type="submit">添加任务</button>
      </form>
    </div>
    <div class="side-stack">
      <section class="card compact-card">
        <div class="section-head">
          <h2>管理员通知</h2>
          <div class="section-note">责任人提交资料后，自动通知此微信联系人。</div>
        </div>
        <form method="post" action="/settings/admin" class="import-grid">
          <label>管理员微信昵称或备注名
            <input name="admin_recipient" required maxlength="100"
              value="{html.escape(admin_recipient, quote=True)}"
              placeholder="例如：邓宇聪">
          </label>
          <button type="submit">保存</button>
        </form>
        <p class="import-tip">需能被机器人小号准确搜索到。</p>
      </section>
      <section class="card compact-card">
        <div class="section-head">
          <h2>Excel 批量导入</h2>
          <div class="section-note">每一行导入为一个任务。</div>
        </div>
        <form method="post" action="/tasks/import" enctype="multipart/form-data" class="import-grid">
          <label>选择 Excel 文件
            <input name="file" type="file" accept=".xlsx,.xlsm" required>
          </label>
          <button type="submit">导入</button>
        </form>
        <p class="import-tip">表头：任务内容及要求、责任人、完成期限。</p>
      </section>
    </div>
  </section>
  <div class="section-head task-list-title">
    <h2>{'待处理任务' if view == 'pending' else '已完成任务'}</h2>
    <div class="section-note">当前页 {len(visible_tasks)} 项，全部 {total_count} 项</div>
  </div>
  <section class="task-list">
    {''.join(cards) or '<div class="card">当前页面没有任务。</div>'}
  </section>
</main>
<script>
async function copySubmissionLink(button) {{
  const original = button.textContent;
  try {{
    await navigator.clipboard.writeText(button.dataset.link);
    button.textContent = "已复制";
  }} catch (error) {{
    window.prompt("请复制提交链接", button.dataset.link);
  }}
  setTimeout(() => button.textContent = original, 1600);
}}
</script>
</body>
</html>"""
    return body.encode("utf-8")


def submit_page(
    task: sqlite3.Row, message: str = "", error: bool = False
) -> bytes:
    notice = ""
    if message:
        css_class = "notice error" if error else "notice"
        notice = f'<div class="{css_class}">{html.escape(message)}</div>'
    body = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>提交任务资料</title>
  <style>{common_styles()}{premium_styles()}</style>
</head>
<body>
<main class="upload-shell">
  <section class="submit-hero">
    <div class="submit-kicker">TASK SUBMISSION</div>
    <h1>任务资料提交</h1>
    <p class="hint">请核对任务信息，并在期限前完成资料上传。</p>
  </section>
  {notice}
  <div class="steps" aria-hidden="true">
    <div class="step active"></div>
    <div class="step active"></div>
    <div class="step"></div>
  </div>
  <section class="card">
    <div class="task-id">任务 #{task['id']}</div>
    <div class="task-summary">
      <div>
        <span class="summary-label">任务内容</span>
        <div class="summary-value">{html.escape(task['content'])}</div>
      </div>
      <div>
        <span class="summary-label">责任人</span>
        <div class="summary-value">{html.escape(task['assignee'])}</div>
      </div>
      <div class="deadline-box">
        <span class="summary-label">提交期限</span>
        <div class="summary-value">{html.escape(task['deadline'])}</div>
      </div>
    </div>
  </section>
  <section class="card">
    <h2>上传提交资料</h2>
    <form method="post" enctype="multipart/form-data">
      <label>提交说明
        <textarea name="note" maxlength="2000" placeholder="填写完成情况、资料说明或需要备注的信息"></textarea>
      </label>
      <br>
      <div class="file-field">
        <label>选择资料
          <input name="file" type="file" required>
        </label>
        <div class="hint" style="margin:8px 0 0">
          支持文档、图片及压缩包，单个文件最大 20 MB。
        </div>
      </div>
      <br>
      <button class="submit-button" type="submit">确认提交资料</button>
    </form>
  </section>
  <p class="privacy-note">此链接仅用于当前任务，请勿转发给无关人员。</p>
</main>
</body>
</html>"""
    return body.encode("utf-8")


def submit_success_page(task: sqlite3.Row, submission: sqlite3.Row) -> bytes:
    body = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>提交成功</title>
  <style>{common_styles()}{premium_styles()}
    .success-panel {{ text-align:center; padding:38px 24px; }}
    .success-icon {{ width:74px; height:74px; margin:0 auto 18px; border-radius:50%;
      display:grid; place-items:center; color:white; font-size:38px; font-weight:800;
      background:linear-gradient(135deg,#08bd63,#07944f);
      box-shadow:0 14px 32px rgba(7,153,79,.24); }}
    .success-title {{ font-size:26px; font-weight:800; margin-bottom:8px; }}
    .success-actions {{ display:flex; justify-content:center; gap:10px; flex-wrap:wrap;
      margin-top:22px; }}
    @media (max-width:800px) {{
      .success-actions {{ display:grid; grid-template-columns:1fr; }}
      .success-actions a {{ width:100%; }}
    }}
  </style>
</head>
<body>
<main class="upload-shell">
  <section class="card success-panel">
    <div class="success-icon">✓</div>
    <div class="success-title">资料提交成功</div>
    <p class="hint">系统已记录你的提交，并通知管理员查看。</p>
    <div class="task-summary" style="text-align:left; margin-top:24px">
      <div>
        <span class="summary-label">任务内容</span>
        <div class="summary-value">{html.escape(task['content'])}</div>
      </div>
      <div>
        <span class="summary-label">提交文件</span>
        <div class="summary-value">{html.escape(submission['original_filename'])}</div>
      </div>
      <div>
        <span class="summary-label">提交时间</span>
        <div class="summary-value">{html.escape(submission['submitted_at'].replace('T', ' '))}</div>
      </div>
    </div>
    <div class="success-actions">
      <a class="button" href="/submit/{task['submit_token']}">返回任务页面</a>
      <a class="button secondary" href="/submit/{task['submit_token']}">继续补充提交</a>
    </div>
  </section>
  <p class="privacy-note">可以关闭此页面，后台状态已自动更新。</p>
</main>
</body>
</html>"""
    return body.encode("utf-8")


def task_detail_page(task: sqlite3.Row) -> bytes:
    rows = []
    for submission in get_submissions(task["id"]):
        rows.append(
            f"""
            <tr>
              <td>{html.escape(submission['submitted_at'].replace('T', ' '))}</td>
              <td>{html.escape(submission['note']) or '—'}</td>
              <td><a href="/submissions/{submission['id']}/download">
                {html.escape(submission['original_filename'])}
              </a></td>
            </tr>
            """
        )
    body = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>任务提交详情</title>
  <style>{common_styles()}{premium_styles()}</style>
</head>
<body>
<main>
  <a href="/">← 返回任务列表</a>
  <h1>任务提交详情</h1>
  <section class="card">
    <p><strong>任务：</strong>{html.escape(task['content'])}</p>
    <p><strong>期限：</strong>{html.escape(task['deadline'])}</p>
    <p><strong>责任人：</strong>{html.escape(task['assignee'])}</p>
  </section>
  <section class="card table-wrap">
    <table>
      <thead><tr><th>提交时间</th><th>提交说明</th><th>资料</th></tr></thead>
      <tbody>{''.join(rows) or '<tr><td colspan="3">尚未提交资料。</td></tr>'}</tbody>
    </table>
  </section>
</main>
</body>
</html>"""
    return body.encode("utf-8")


class Handler(BaseHTTPRequestHandler):
    def send_html(self, content: bytes, status: int = 200) -> None:
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def send_text_error(self, status: int, message: str) -> None:
        content = message.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def redirect(self, message: str, error: bool = False) -> None:
        query = urllib.parse.urlencode(
            {"message": message, "error": "1" if error else "0"}
        )
        self.send_response(303)
        self.send_header("Location", f"/?{query}")
        self.end_headers()

    def redirect_to(self, path: str, message: str, error: bool = False) -> None:
        separator = "&" if "?" in path else "?"
        query = urllib.parse.urlencode(
            {"message": message, "error": "1" if error else "0"}
        )
        self.send_response(303)
        self.send_header("Location", f"{path}{separator}{query}")
        self.end_headers()

    def read_form(self) -> dict[str, str]:
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length).decode("utf-8")
        parsed = urllib.parse.parse_qs(raw, keep_blank_values=True)
        return {key: values[0].strip() for key, values in parsed.items()}

    def request_base_url(self) -> str:
        proto = self.headers.get("X-Forwarded-Proto", "http").split(",")[0].strip()
        host = self.headers.get("X-Forwarded-Host", self.headers.get("Host", ""))
        return f"{proto}://{host}"

    def read_multipart_parts(self) -> dict[str, tuple[str, object]]:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0 or length > MAX_UPLOAD_SIZE + 1024 * 1024:
            raise ValueError("上传内容为空或超过 20 MB。")
        content_type = self.headers.get("Content-Type", "")
        if not content_type.startswith("multipart/form-data"):
            raise ValueError("提交格式不正确。")
        raw = self.rfile.read(length)
        message = BytesParser(policy=default).parsebytes(
            (
                f"Content-Type: {content_type}\r\n"
                "MIME-Version: 1.0\r\n\r\n"
            ).encode("ascii")
            + raw
        )
        parts: dict[str, tuple[str, object]] = {}
        for part in message.iter_parts():
            disposition = part.get("Content-Disposition", "")
            if "form-data" not in disposition:
                continue
            name = part.get_param("name", header="Content-Disposition")
            if not name:
                continue
            filename = part.get_filename() or ""
            if filename:
                parts[name] = (filename, part.get_payload(decode=True) or b"")
            else:
                parts[name] = ("", part.get_content().strip())
        return parts

    def read_submission_multipart(self) -> tuple[str, str, bytes]:
        parts = self.read_multipart_parts()
        note = str(parts.get("note", ("", ""))[1])[:2000]
        filename, payload = parts.get("file", ("", b""))
        file_data = payload if isinstance(payload, bytes) else b""
        if not filename or not file_data:
            raise ValueError("请选择要上传的资料。")
        if len(file_data) > MAX_UPLOAD_SIZE:
            raise ValueError("文件超过 20 MB。")
        return note, filename, file_data

    def read_excel_upload(self) -> tuple[str, bytes]:
        parts = self.read_multipart_parts()
        filename, payload = parts.get("file", ("", b""))
        file_data = payload if isinstance(payload, bytes) else b""
        if not filename or not file_data:
            raise ValueError("请选择要导入的 Excel 文件。")
        suffix = Path(filename).suffix.lower()
        if suffix not in {".xlsx", ".xlsm"}:
            raise ValueError("请上传 .xlsx 或 .xlsm 文件。")
        return filename, file_data

    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/":
            query = urllib.parse.parse_qs(parsed.query)
            self.send_html(
                page(
                    self.request_base_url(),
                    query.get("view", ["pending"])[0],
                    query.get("message", [""])[0],
                    query.get("error", ["0"])[0] == "1",
                )
            )
            return
        parts = parsed.path.strip("/").split("/")
        if len(parts) == 2 and parts[0] == "submit":
            task = get_task_by_token(parts[1])
            if task is None:
                self.send_text_error(404, "提交链接不存在")
                return
            query = urllib.parse.parse_qs(parsed.query)
            self.send_html(
                submit_page(
                    task,
                    query.get("message", [""])[0],
                    query.get("error", ["0"])[0] == "1",
                )
            )
            return
        if len(parts) == 3 and parts[0] == "submit" and parts[2] == "success":
            task = get_task_by_token(parts[1])
            if task is None:
                self.send_text_error(404, "提交链接不存在")
                return
            query = urllib.parse.parse_qs(parsed.query)
            try:
                submission_id = int(query.get("submission_id", ["0"])[0])
            except ValueError:
                submission_id = 0
            submission = get_submission(submission_id)
            if submission is None or submission["task_id"] != task["id"]:
                self.send_text_error(404, "提交记录不存在")
                return
            self.send_html(submit_success_page(task, submission))
            return
        if len(parts) == 2 and parts[0] == "tasks":
            try:
                task = get_task(int(parts[1]))
            except ValueError:
                task = None
            if task is None:
                self.send_error(404)
                return
            self.send_html(task_detail_page(task))
            return
        if len(parts) == 3 and parts[0] == "submissions" and parts[2] == "download":
            try:
                submission = get_submission(int(parts[1]))
            except ValueError:
                submission = None
            if submission is None:
                self.send_error(404)
                return
            file_path = UPLOAD_DIR / submission["stored_filename"]
            if not file_path.is_file():
                self.send_text_error(404, "文件不存在")
                return
            data = file_path.read_bytes()
            content_type = mimetypes.guess_type(submission["original_filename"])[0]
            self.send_response(200)
            self.send_header("Content-Type", content_type or "application/octet-stream")
            encoded_name = urllib.parse.quote(submission["original_filename"])
            self.send_header(
                "Content-Disposition", f"attachment; filename*=UTF-8''{encoded_name}"
            )
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if parsed.path == "/api/health":
            data = json.dumps(
                {
                    "ok": True,
                    "normal_reminder_slots": NORMAL_REMINDER_SLOTS,
                    "urgent_reminder_slots": URGENT_REMINDER_SLOTS,
                },
                ensure_ascii=False,
            ).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        self.send_error(404)

    def do_POST(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        parts = parsed.path.strip("/").split("/")
        if len(parts) == 2 and parts[0] == "submit":
            task = get_task_by_token(parts[1])
            if task is None:
                self.send_text_error(404, "提交链接不存在")
                return
            try:
                note, filename, file_data = self.read_submission_multipart()
                submission = save_submission(task["id"], note, filename, file_data)
            except ValueError as exc:
                self.redirect_to(parsed.path, str(exc), True)
                return
            admin_recipient = get_admin_recipient()
            if admin_recipient:
                ok, detail = send_admin_submission(admin_recipient, task, submission)
                print(
                    f"[ADMIN_NOTIFY] task={task['id']} admin={admin_recipient} "
                    f"success={ok} detail={detail}"
                )
            self.send_response(303)
            self.send_header(
                "Location",
                f"{parsed.path}/success?submission_id={submission['id']}",
            )
            self.end_headers()
            return

        if parsed.path == "/settings/admin":
            form = self.read_form()
            admin_recipient = form.get("admin_recipient", "").strip()
            if not admin_recipient:
                self.redirect("管理员微信昵称不能为空。", True)
                return
            set_admin_recipient(admin_recipient)
            self.redirect(f"管理员通知对象已设置为：{admin_recipient}")
            return

        if parsed.path == "/tasks/remind-all":
            tasks = [task for task in get_tasks() if not task["submission_count"]]
            success_count = 0
            failed: list[str] = []
            for task in tasks:
                link = submission_url(task, self.request_base_url())
                ok, detail = send_wechat_message(
                    task["assignee"],
                    reminder_message(task, link, is_urgent_or_overdue(task)),
                )
                if ok:
                    success_count += 1
                else:
                    failed.append(f"{task['assignee']}：{detail}")
            message = f"一键提醒完成：成功 {success_count} 个，失败 {len(failed)} 个。"
            if failed:
                message += " " + "；".join(failed[:3])
            self.redirect(message, bool(failed))
            return

        if parsed.path == "/tasks":
            form = self.read_form()
            content = form.get("content", "")
            deadline = form.get("deadline", "")
            assignee = form.get("assignee", "")
            if not all((content, deadline, assignee)):
                self.redirect("任务内容、期限和责任人都不能为空。", True)
                return
            create_task(content, deadline.replace("T", " "), assignee)
            self.redirect("任务已添加。")
            return

        if parsed.path == "/tasks/import":
            try:
                filename, file_data = self.read_excel_upload()
                tasks, errors = parse_task_excel(file_data)
                imported = create_tasks_bulk(tasks)
            except ValueError as exc:
                self.redirect(f"导入失败：{exc}", True)
                return
            message = f"已从 {filename} 导入 {imported} 个任务。"
            if errors:
                preview = "；".join(errors[:3])
                if len(errors) > 3:
                    preview += f"；另有 {len(errors) - 3} 行错误"
                message += f" 跳过 {len(errors)} 行：{preview}"
            self.redirect(message)
            return

        if len(parts) == 3 and parts[0] == "tasks" and parts[2] == "notify":
            try:
                task_id = int(parts[1])
            except ValueError:
                self.send_error(400)
                return
            task = get_task(task_id)
            if task is None:
                self.send_error(404)
                return
            link = submission_url(task, self.request_base_url())
            message = notification_message(task, link)
            ok, detail = send_wechat_message(task["assignee"], message)
            if ok:
                mark_notified(task_id)
            self.redirect(detail, not ok)
            return

        if len(parts) == 3 and parts[0] == "tasks" and parts[2] == "delete":
            try:
                task_id = int(parts[1])
            except ValueError:
                self.send_error(400)
                return
            deleted = delete_task(task_id)
            self.redirect("任务已删除。" if deleted else "任务不存在。", not deleted)
            return

        self.send_error(404)

    def log_message(self, format: str, *args: object) -> None:
        print(f"[HTTP] {format % args}")


def main() -> None:
    initialize_database()
    reminder_thread = threading.Thread(
        target=reminder_loop,
        name="daily-reminder",
        daemon=True,
    )
    reminder_thread.start()
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"微信任务跟踪后台已启动：http://{HOST}:{PORT}")
    print(f"常规提醒：{', '.join(NORMAL_REMINDER_SLOTS)}")
    print(f"临期/逾期提醒：{', '.join(URGENT_REMINDER_SLOTS)}")
    print("按 Ctrl+C 停止。")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
